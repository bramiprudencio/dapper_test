import pandas as pd
import io
from openpyxl import load_workbook
from .base import BaseScraper 

class ColombiaScraper(BaseScraper):
    def __init__(self):
        super().__init__() # IMPORTANTE: Inicializar el padre para tener self.session
        self.pais = "Colombia"
        self.url_api = "https://www.camara.gov.co/wp-admin/admin-ajax.php"
        self.payload = {
            'action': 'download_proyectos_ley_xlsx',
            '_ajax_nonce': '376dd2bad5',
            'tipo': 'All', 'estado': 'All', 'origen': 'All', 
            'legislatura': 'All', 'comision_adv': 'All'
        }

    def scrape(self):
        print(f"🇨🇴 Iniciando scraping de {self.pais} via Excel en Memoria...")
        
        try:
            # 1. Descargar a RAM usando la sesión (mantiene cookies y headers)
            response = self.session.post(self.url_api, data=self.payload, timeout=30)
            response.raise_for_status() # Lanza error si es 404 o 500
            
            # Convertimos los bytes descargados en un "archivo virtual"
            archivo_en_memoria = io.BytesIO(response.content)

            # 2. Cargar DataFrame con Pandas (para leer datos y encabezados correctamente)
            # read_excel entiende io.BytesIO perfectamente
            df = pd.read_excel(archivo_en_memoria)
            
            # --- Corrección de nombres de columnas ---
            # A veces vienen con espacios extra: "Título " -> "Título"
            df.columns = df.columns.str.strip()

            # 3. Cargar OpenPyXL (para extraer hipervínculos)
            # Necesitamos resetear el puntero del archivo virtual o crear uno nuevo
            archivo_en_memoria.seek(0) 
            wb = load_workbook(filename=archivo_en_memoria)
            ws = wb.active
            
            # 4. Extracción de Links
            links = []
            
            # openpyxl itera filas físicas. [1:] salta el encabezado.
            # Asumimos columna P (índice 15).
            # Verificamos que no nos pasemos de la cantidad de datos del DF
            celdas_links = ws['P'][1 : len(df) + 1]
            
            for cell in celdas_links:
                if cell.hyperlink:
                    links.append(cell.hyperlink.target)
                else:
                    # Intento secundario: Si hay texto "http" en la celda
                    val = str(cell.value) if cell.value else ""
                    links.append(val if "http" in val else None)
            
            # Asegurar longitud
            while len(links) < len(df):
                links.append(None)
                
            # Asignar al DataFrame temporalmente para facilitar el loop
            df['Link_Extraido'] = links
            
            # Limpieza de fechas en Pandas
            if 'Fecha Cámara' in df.columns:
                df['Fecha Cámara'] = pd.to_datetime(df['Fecha Cámara'], errors='coerce')
            
            data_limpia = []

            # 5. Construir lista de diccionarios
            for index, row in df.iterrows():
                # Usamos .get() para ser tolerantes a cambios de nombre leves
                titulo = row.get('Título')
                fecha = row.get('Fecha Cámara')
                resumen = row.get('Objeto del proyecto')
                estado = row.get('Estado de Ley')
                
                # Conversión segura de NaT/NaN a None para base de datos
                if pd.isna(fecha): fecha = None
                if pd.isna(titulo): titulo = None
                
                if titulo:
                    data_limpia.append({
                        "titulo": str(titulo).strip(),
                        "fecha_radicacion": fecha,
                        "resumen": str(resumen) if resumen else None,
                        "enlace": row['Link_Extraido'],
                        "estado": str(estado) if estado else None,
                        "pais": self.pais
                    })
            
            print(f"✅ Se obtuvieron {len(data_limpia)} registros limpios.")
            return data_limpia

        except Exception as e:
            print(f"❌ Error descargando/procesando Excel: {e}")
            return [] # Retornar lista vacía para no romper el main loop